# -*- coding: utf-8 -*-
"""
虚拟净值数据提取系统 - 单文件版
功能：读取中间表，匹配估值表，提取数据，写入模板
"""
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkcalendar import Calendar
import threading
import os
import re
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter
import xlrd

# ==================== 日志配置 ====================
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ==================== 数据类定义 ====================
@dataclass
class MiddleTableRow:
    """中间表数据行"""
    product_code: str
    account_set: str
    cert_id: str
    product_name: str
    account_code: str
    account_name: str
    match_word: str


@dataclass
class ValuationData:
    """估值数据"""
    cert_id: str
    virtual_net_value: str


# ==================== Excel读取器 ====================
class ExcelReader:
    """Excel读取器 - 支持xls和xlsx格式"""
    
    @staticmethod
    def read_middle_table(file_path: str) -> List[MiddleTableRow]:
        """读取中间表"""
        rows = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb['中间表']
            
            # 找到表头行
            header_row = None
            header_map = {}
            expected_headers = ['产品代码', '帐套编号', '证件号', '产品名称', '专户代码', '专户名称', '匹配词']
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row and row[0] in expected_headers:
                    header_row = row_idx
                    for col_idx, header in enumerate(row):
                        if header:
                            header_map[header] = col_idx
                    break
            
            if header_row is None:
                logger.error("中间表：未找到表头行")
                return rows
            
            # 读取数据行
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or row[0] is None:
                    continue
                
                # 检查是否含有#NA
                if any('#NA' in str(cell) if cell else False for cell in row):
                    logger.warning(f"跳过含有#NA的行")
                    continue
                
                try:
                    product_code_col = header_map.get('产品代码', 0)
                    account_set_col = header_map.get('帐套编号', 1)
                    cert_id_col = header_map.get('证件号', 2)
                    product_name_col = header_map.get('产品名称', 3)
                    account_code_col = header_map.get('专户代码', 4)
                    account_name_col = header_map.get('专户名称', 5)
                    match_word_col = header_map.get('匹配词', 6)
                    
                    middle_row = MiddleTableRow(
                        product_code=str(row[product_code_col]).strip() if row[product_code_col] else '',
                        account_set=str(row[account_set_col]).strip() if row[account_set_col] else '',
                        cert_id=str(row[cert_id_col]).strip() if row[cert_id_col] else '',
                        product_name=str(row[product_name_col]).strip() if row[product_name_col] else '',
                        account_code=str(row[account_code_col]).strip() if row[account_code_col] else '',
                        account_name=str(row[account_name_col]).strip() if row[account_name_col] else '',
                        match_word=str(row[match_word_col]).strip() if row[match_word_col] else ''
                    )
                    rows.append(middle_row)
                except Exception as e:
                    logger.error(f"解析中间表行失败: {e}")
            
            wb.close()
        except Exception as e:
            logger.error(f"读取中间表失败: {e}")
        
        return rows
    
    @staticmethod
    def read_valuation_file(file_path: str, match_words: List[str]) -> Tuple[List[ValuationData], bool]:
        """读取估值表文件"""
        data = []
        found_match_word = False
        
        try:
            if file_path.endswith('.xls'):
                workbook = xlrd.open_workbook(file_path, encoding_override='utf-8')
                for sheet in workbook.sheets():
                    result, found = ExcelReader._extract_from_sheet_xls(sheet, match_words)
                    data.extend(result)
                    found_match_word = found_match_word or found
            else:  # xlsx
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    result, found = ExcelReader._extract_from_sheet_xlsx(ws, match_words)
                    data.extend(result)
                    found_match_word = found_match_word or found
                wb.close()
        except Exception as e:
            logger.error(f"读取估值文件失败 {file_path}: {e}")
            return [], False
        
        return data, found_match_word
    
    @staticmethod
    def _extract_from_sheet_xls(sheet, match_words: List[str]) -> Tuple[List[ValuationData], bool]:
        """从XLS sheet中提取数据"""
        data = []
        found_match_word = False
        
        try:
            # 查找虚拟净值列和证件号列
            virtual_net_value_col = None
            cert_id_col = None
            header_row = None
            
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell_value = str(sheet.cell_value(row_idx, col_idx)).strip()
                    if '虚拟净值' in cell_value or '虚拟单位净值' in cell_value:
                        virtual_net_value_col = col_idx
                        header_row = row_idx
                    if cell_value == '证件号':
                        cert_id_col = col_idx
                        if header_row is None:
                            header_row = row_idx
                
                if virtual_net_value_col is not None:
                    break
            
            if virtual_net_value_col is None:
                logger.warning(f"Sheet {sheet.name}: 未找到虚拟净值列")
                return data, False
            
            # 提取数据
            for row_idx in range(header_row + 1, sheet.nrows):
                cert_id = str(sheet.cell_value(row_idx, cert_id_col)).strip() if cert_id_col is not None else ''
                net_value = str(sheet.cell_value(row_idx, virtual_net_value_col)).strip()
                
                if net_value and net_value != 'nan' and cert_id and cert_id != 'nan':
                    for match_word in match_words:
                        if match_word in sheet.name:
                            found_match_word = True
                            break
                    
                    data.append(ValuationData(
                        cert_id=cert_id,
                        virtual_net_value=net_value
                    ))
        except Exception as e:
            logger.error(f"解析XLS sheet失败: {e}")
        
        return data, found_match_word
    
    @staticmethod
    def _extract_from_sheet_xlsx(ws, match_words: List[str]) -> Tuple[List[ValuationData], bool]:
        """从XLSX sheet中提取数据"""
        data = []
        found_match_word = False
        
        try:
            # 查找虚拟净值列和证件号列
            virtual_net_value_col = None
            cert_id_col = None
            header_row = None
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
                for col_idx, cell in enumerate(row):
                    cell_value = str(cell.value).strip() if cell.value else ''
                    if '虚拟净值' in cell_value or '虚拟单位净值' in cell_value:
                        virtual_net_value_col = col_idx
                        header_row = row_idx
                    if cell_value == '证件号':
                        cert_id_col = col_idx
                        if header_row is None:
                            header_row = row_idx
                
                if virtual_net_value_col is not None:
                    break
            
            if virtual_net_value_col is None:
                logger.warning(f"Sheet {ws.title}: 未找到虚拟净值列")
                return data, False
            
            # 提取数据
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx <= header_row:
                    continue
                
                cert_id = str(row[cert_id_col]).strip() if cert_id_col is not None and cert_id_col < len(row) and row[cert_id_col] else ''
                net_value = str(row[virtual_net_value_col]).strip() if virtual_net_value_col < len(row) and row[virtual_net_value_col] else ''
                
                if net_value and net_value != 'None' and cert_id and cert_id != 'None':
                    for match_word in match_words:
                        if match_word in ws.title:
                            found_match_word = True
                            break
                    
                    data.append(ValuationData(
                        cert_id=cert_id,
                        virtual_net_value=net_value
                    ))
        except Exception as e:
            logger.error(f"解析XLSX sheet失败: {e}")
        
        return data, found_match_word
    
    @staticmethod
    def find_matching_file(excel_folder: str, match_word: str, date_formats: List[str]) -> Optional[str]:
        """查找匹配的估值表文件"""
        excel_files = []
        
        try:
            for file in os.listdir(excel_folder):
                if file.endswith(('.xls', '.xlsx')):
                    excel_files.append(os.path.join(excel_folder, file))
        except Exception as e:
            logger.error(f"读取Excel文件夹失败: {e}")
            return None
        
        # 在文件名中查找匹配词
        for file_path in excel_files:
            file_name = os.path.basename(file_path)
            if match_word in file_name:
                return file_path
        
        # 如果没找到，尝试更灵活的匹配
        for file_path in excel_files:
            file_name = os.path.basename(file_path)
            if match_word.lower() in file_name.lower():
                return file_path
        
        return None


# ==================== 数据处理器 ====================
class DataProcessor:
    """数据处理器"""
    
    @staticmethod
    def match_valuation_data(middle_table: List[MiddleTableRow], 
                            excel_folder: str, 
                            date_formats: List[str],
                            progress_callback=None) -> Tuple[List[Dict], List[str]]:
        """匹配并处理估值数据"""
        all_data = []
        errors = []
        matched_match_words = set()
        
        for idx, middle_row in enumerate(middle_table):
            if progress_callback:
                progress_callback(idx, len(middle_table))
            
            match_word = middle_row.match_word
            if not match_word:
                errors.append(f"产品 {middle_row.product_name} 没有匹配词")
                continue
            
            # 查找对应的估值文件
            excel_file = ExcelReader.find_matching_file(excel_folder, match_word, date_formats)
            
            if not excel_file:
                errors.append(f"未找到匹配词 '{match_word}' 对应的估值表文件")
                continue
            
            # 读取估值文件
            valuation_data, found_match = ExcelReader.read_valuation_file(excel_file, [match_word])
            
            if not found_match:
                errors.append(f"估值文件 {os.path.basename(excel_file)} 中未找到匹配词 '{match_word}'")
            
            if not valuation_data:
                errors.append(f"从文件 {os.path.basename(excel_file)} 中未提取到数据")
                continue
            
            matched_match_words.add(match_word)
            
            # 匹配证件号
            for data in valuation_data:
                if data.cert_id == middle_row.cert_id:
                    all_data.append({
                        'account_set': middle_row.account_set,
                        'account_code': middle_row.account_code,
                        'account_name': middle_row.account_name,
                        'cert_id': middle_row.cert_id,
                        'virtual_net_value': data.virtual_net_value,
                        'match_word': match_word
                    })
        
        # 完整性校验
        middle_cert_ids = set(row.cert_id for row in middle_table if row.match_word)
        extracted_cert_ids = set(row['cert_id'] for row in all_data)
        
        missing_cert_ids = middle_cert_ids - extracted_cert_ids
        if missing_cert_ids:
            for cert_id in missing_cert_ids:
                matching_row = next((r for r in middle_table if r.cert_id == cert_id), None)
                if matching_row:
                    errors.append(f"中间表中的证件号 {cert_id} 在估值表中未找到")
        
        return all_data, errors
    
    @staticmethod
    def validate_data(middle_table: List[MiddleTableRow], 
                     valuation_data: List[Dict]) -> List[str]:
        """数据完整性校验"""
        errors = []
        
        middle_set = set((row.account_set, row.account_code, row.cert_id) 
                        for row in middle_table if row.match_word)
        valuation_set = set((row['account_set'], row['account_code'], row['cert_id']) 
                           for row in valuation_data)
        
        missing = middle_set - valuation_set
        if missing:
            for item in missing:
                errors.append(f"缺少数据: 账套={item[0]}, 专户代码={item[1]}, 证件号={item[2]}")
        
        return errors


# ==================== Excel写入器 ====================
class ExcelWriter:
    """Excel写入器 - 使用openpyxl"""
    
    @staticmethod
    def find_header_columns(ws, header_names: List[str]) -> Dict[str, int]:
        """查找表头列号"""
        header_cols = {}
        
        # 遍历所有行查找表头
        for row in ws.iter_rows(max_row=10, values_only=False):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value in header_names:
                        header_cols[cell_value] = col_idx
            
            # 如果找到了所有表头，就停止
            if len(header_cols) == len(header_names):
                break
        
        return header_cols
    
    @staticmethod
    def write_to_template_old(template_path: str, 
                             output_path: str,
                             data: List[Dict],
                             query_date: str) -> bool:
        """写入旧版本模板"""
        try:
            # 打开模板
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 查找表头
            header_names = ['行情日期', '账套编号', '证券代码', '收盘价', '证券品种', '市场']
            header_cols = ExcelWriter.find_header_columns(ws, header_names)
            
            if not all(h in header_cols for h in header_names):
                logger.error(f"旧版本模板：未找到所有表头")
                return False
            
            # 写入数据
            row_idx = 2
            for item in data:
                try:
                    ws.cell(row_idx, header_cols['行情日期']).value = query_date
                    ws.cell(row_idx, header_cols['账套编号']).value = item['account_set']
                    ws.cell(row_idx, header_cols['证券代码']).value = item['account_code']
                    ws.cell(row_idx, header_cols['收盘价']).value = item['virtual_net_value']
                    ws.cell(row_idx, header_cols['证券品种']).value = '资产管理产品'
                    ws.cell(row_idx, header_cols['市场']).value = '场外'
                    
                    # 设置格式为文本
                    for header in header_names:
                        ws.cell(row_idx, header_cols[header]).number_format = '@'
                    
                    row_idx += 1
                except Exception as e:
                    logger.error(f"写入行失败: {e}")
                    continue
            
            # 保存
            wb.save(output_path)
            wb.close()
            
            logger.info(f"成功写入旧版本模板: {output_path}")
            return True
        except Exception as e:
            logger.error(f"写入旧版本模板失败: {e}")
            return False
    
    @staticmethod
    def write_to_template_new(template_path: str,
                             output_path: str,
                             data: List[Dict],
                             query_date: str) -> bool:
        """写入新版本模板"""
        try:
            # 打开模板
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 查找表头
            header_names = ['行情日期', '账套编号', '证券代码', '虚拟单位净值', '证券品种', '市场']
            header_cols = ExcelWriter.find_header_columns(ws, header_names)
            
            if not all(h in header_cols for h in header_names):
                logger.error(f"新版本模板：未找到所有表头")
                return False
            
            # 写入数据
            row_idx = 2
            for item in data:
                try:
                    ws.cell(row_idx, header_cols['行情日期']).value = query_date
                    ws.cell(row_idx, header_cols['账套编号']).value = item['account_set']
                    ws.cell(row_idx, header_cols['证券代码']).value = item['account_code']
                    ws.cell(row_idx, header_cols['虚拟单位净值']).value = item['virtual_net_value']
                    ws.cell(row_idx, header_cols['证券品种']).value = '资产管理产品'
                    ws.cell(row_idx, header_cols['市场']).value = '场外'
                    
                    # 设置格式为文本
                    for header in header_names:
                        ws.cell(row_idx, header_cols[header]).number_format = '@'
                    
                    row_idx += 1
                except Exception as e:
                    logger.error(f"写入行失败: {e}")
                    continue
            
            # 保存
            wb.save(output_path)
            wb.close()
            
            logger.info(f"成功写入新版本模板: {output_path}")
            return True
        except Exception as e:
            logger.error(f"写入新版本模板失败: {e}")
            return False


# ==================== 日期格式化器 ====================
class DateFormatter:
    """日期格式化器"""
    
    @staticmethod
    def format_date(date_obj) -> List[str]:
        """将日期对象转换为多种格式"""
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            except:
                return []
        
        formats = [
            date_obj.strftime('%Y%m%d'),      # YYYYMMDD
            date_obj.strftime('%Y-%m-%d'),    # YYYY-MM-DD
            date_obj.strftime('%Y.%m.%d'),    # YYYY.MM.DD
            date_obj.strftime('%Y年%m月%d日')  # YYYY年MM月DD日
        ]
        
        return formats
    
    @staticmethod
    def get_yyyymmdd(date_obj) -> str:
        """获取YYYYMMDD格式"""
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            except:
                return ''
        
        return date_obj.strftime('%Y%m%d')


# ==================== 主处理函数 ====================
def process_valuation_data(excel_folder: str,
                          query_date: str,
                          valuation_date,
                          progress_callback=None,
                          log_callback=None) -> bool:
    """主处理函数"""
    
    try:
        # 获取脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        middle_table_path = os.path.join(script_dir, '无敌虚拟净值中间表.xlsx')
        
        if not os.path.exists(middle_table_path):
            error_msg = f"中间表不存在: {middle_table_path}"
            if log_callback:
                log_callback(error_msg, "error")
            logger.error(error_msg)
            return False
        
        # 读取中间表
        if log_callback:
            log_callback("正在读取中间表...", "info")
        middle_table = ExcelReader.read_middle_table(middle_table_path)
        
        if not middle_table:
            error_msg = "中间表为空"
            if log_callback:
                log_callback(error_msg, "error")
            logger.error(error_msg)
            return False
        
        if log_callback:
            log_callback(f"成功读取中间表，共 {len(middle_table)} 行数据", "success")
        
        # 获取日期格式
        date_formats = DateFormatter.format_date(valuation_date)
        query_date_yyyymmdd = DateFormatter.get_yyyymmdd(query_date)
        
        # 匹配并处理估值数据
        if log_callback:
            log_callback("正在匹配估值表...", "info")
        valuation_data, errors = DataProcessor.match_valuation_data(
            middle_table, excel_folder, date_formats, progress_callback
        )
        
        # 输出错误信息
        if errors:
            if log_callback:
                log_callback("\n=== 警告和错误 ===", "warning")
                for error in errors:
                    log_callback(f"⚠️  {error}", "warning")
        
        if not valuation_data:
            error_msg = "未提取到任何估值数据"
            if log_callback:
                log_callback(error_msg, "error")
            logger.error(error_msg)
            return False
        
        if log_callback:
            log_callback(f"成功提取 {len(valuation_data)} 条数据", "success")
        
        # 完整性校验
        validation_errors = DataProcessor.validate_data(middle_table, valuation_data)
        if validation_errors:
            if log_callback:
                log_callback("\n=== 数据完整性校验 ===", "warning")
                for error in validation_errors:
                    log_callback(f"❌ {error}", "warning")
        else:
            if log_callback:
                log_callback("✅ 数据完整性校验通过", "success")
        
        # 写入数据
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 旧版本模板
        old_template = os.path.join(script_dir, '手工行情导入模板20251124.xlsx')
        if os.path.exists(old_template):
            old_output = os.path.join(script_dir, f'手工行情导入模板{query_date_yyyymmdd}.xlsx')
            if log_callback:
                log_callback(f"正在写入旧版本模板...", "info")
            success = ExcelWriter.write_to_template_old(old_template, old_output, valuation_data, query_date_yyyymmdd)
            if success and log_callback:
                log_callback(f"✅ 旧版本模板写入成功: {os.path.basename(old_output)}", "success")
        else:
            if log_callback:
                log_callback(f"⚠️  旧版本模板不存在: {old_template}", "warning")
        
        # 新版本模板
        new_template = os.path.join(script_dir, '手工行情导入模板5月1日后新版本.xlsx')
        if os.path.exists(new_template):
            new_output = os.path.join(script_dir, f'【新版本】手工行情导入模板{query_date_yyyymmdd}.xlsx')
            if log_callback:
                log_callback(f"正在写入新版本模板...", "info")
            success = ExcelWriter.write_to_template_new(new_template, new_output, valuation_data, query_date_yyyymmdd)
            if success and log_callback:
                log_callback(f"✅ 新版本模板写入成功: {os.path.basename(new_output)}", "success")
        else:
            if log_callback:
                log_callback(f"⚠️  新版本模板不存在: {new_template}", "warning")
        
        if log_callback:
            log_callback("\n✅ 数据处理完成！", "success")
        
        return True
    
    except Exception as e:
        error_msg = f"处理过程中出现错误: {e}"
        if log_callback:
            log_callback(error_msg, "error")
        logger.error(error_msg)
        return False


# ==================== GUI 界面 ====================
class ValuationGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("虚拟净值提取系统")
        self.root.geometry("1000x800")
        
        self.excel_folder = tk.StringVar()
        self.query_date = tk.StringVar()
        self.valuation_date = None
        
        self._create_widgets()
    
    def _create_widgets(self):
        """创建UI组件"""
        
        # 顶部标题
        title_label = tk.Label(self.root, text="虚拟净值数据提取系统", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 第一行：Excel文件夹选择
        folder_frame = tk.Frame(self.root)
        folder_frame.pack(padx=10, pady=5, fill=tk.X)
        
        tk.Label(folder_frame, text="Excel文件夹:", font=("Arial", 10)).pack(side=tk.LEFT)
        tk.Entry(folder_frame, textvariable=self.excel_folder, width=50, state='readonly').pack(side=tk.LEFT, padx=5)
        tk.Button(folder_frame, text="浏览...", command=self._select_folder).pack(side=tk.LEFT)
        
        # 第二行：行情日期选择
        date_frame = tk.Frame(self.root)
        date_frame.pack(padx=10, pady=5, fill=tk.X)
        
        tk.Label(date_frame, text="行情日期(YYYYMMDD):", font=("Arial", 10)).pack(side=tk.LEFT)
        tk.Entry(date_frame, textvariable=self.query_date, width=20).pack(side=tk.LEFT, padx=5)
        tk.Button(date_frame, text="帮助", command=self._help_date_format).pack(side=tk.LEFT)
        
        # 第三行：估值表时间选择
        valuation_frame = tk.Frame(self.root)
        valuation_frame.pack(padx=10, pady=5, fill=tk.X)
        
        tk.Label(valuation_frame, text="估值表时间:", font=("Arial", 10)).pack(side=tk.LEFT)
        self.valuation_date_label = tk.Label(valuation_frame, text="未选择", font=("Arial", 10), fg="red")
        self.valuation_date_label.pack(side=tk.LEFT, padx=5)
        tk.Button(valuation_frame, text="选择日期...", command=self._select_valuation_date).pack(side=tk.LEFT)
        
        # 进度条
        progress_frame = tk.Frame(self.root)
        progress_frame.pack(padx=10, pady=10, fill=tk.X)
        
        tk.Label(progress_frame, text="处理进度:", font=("Arial", 10)).pack(side=tk.LEFT)
        self.progress_label = tk.Label(progress_frame, text="0/0", font=("Arial", 10))
        self.progress_label.pack(side=tk.LEFT, padx=5)
        
        self.progress = tk.Canvas(progress_frame, height=20, bg='white', relief=tk.SUNKEN, bd=1)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # 日志窗口
        log_frame = tk.LabelFrame(self.root, text="日志", font=("Arial", 10, "bold"))
        log_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, width=80, font=("Courier", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 配置日志文本的标签样式
        self.log_text.tag_config("error", foreground="red")
        self.log_text.tag_config("warning", foreground="orange")
        self.log_text.tag_config("success", foreground="green")
        self.log_text.tag_config("info", foreground="black")
        
        # 底部按钮
        button_frame = tk.Frame(self.root)
        button_frame.pack(padx=10, pady=10, fill=tk.X)
        
        tk.Button(button_frame, text="开始处理", command=self._start_process, 
                 bg="green", fg="white", font=("Arial", 11, "bold"), width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="清空日志", command=self._clear_log, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="退出", command=self.root.quit, width=15).pack(side=tk.RIGHT, padx=5)
    
    def _select_folder(self):
        """选择Excel文件夹"""
        folder = filedialog.askdirectory(title="选择Excel文件夹")
        if folder:
            self.excel_folder.set(folder)
            self._log(f"✅ 选择文件夹: {folder}", "success")
    
    def _select_valuation_date(self):
        """选择估值表日期"""
        # 创建日期选择窗口
        date_window = tk.Toplevel(self.root)
        date_window.title("选择估值表日期")
        date_window.geometry("400x350")
        
        cal = Calendar(date_window, selectmode='day')
        cal.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        
        def confirm_date():
            selected_date = cal.get_date()
            self.valuation_date = datetime.strptime(selected_date, '%m/%d/%y')
            
            # 显示多种格式
            date_formats = DateFormatter.format_date(self.valuation_date)
            self.valuation_date_label.config(
                text=f"{date_formats[0]} | {date_formats[1]} | {date_formats[2]} | {date_formats[3]}",
                fg="green"
            )
            
            self._log(f"✅ 选择估值表日期: {date_formats[0]}", "success")
            date_window.destroy()
        
        tk.Button(date_window, text="确认", command=confirm_date, 
                 bg="blue", fg="white", font=("Arial", 11)).pack(pady=10)
    
    def _help_date_format(self):
        """帮助信息"""
        messagebox.showinfo("日期格式说明", 
                          "请输入YYYYMMDD格式的日期，例如：20251124\n\n"
                          "其中：\nYYYY - 年份（4位数字）\nMM - 月份���2位数字）\nDD - 日期（2位数字）")
    
    def _log(self, message, tag="info"):
        """写入日志"""
        self.log_text.insert(tk.END, message + "\n", tag)
        self.log_text.see(tk.END)
        self.root.update()
    
    def _clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
    
    def _update_progress(self, current, total):
        """更新进度条"""
        self.progress_label.config(text=f"{current}/{total}")
        if total > 0:
            progress_percent = current / total
            self.progress.delete("all")
            self.progress.create_rectangle(0, 0, self.progress.winfo_width() * progress_percent, 
                                         self.progress.winfo_height(), fill="blue")
        self.root.update()
    
    def _validate_inputs(self) -> bool:
        """验证输入"""
        if not self.excel_folder.get():
            messagebox.showerror("错误", "请选择Excel文件夹")
            return False
        
        if not os.path.exists(self.excel_folder.get()):
            messagebox.showerror("错误", f"文件夹不存在: {self.excel_folder.get()}")
            return False
        
        if not self.query_date.get():
            messagebox.showerror("错误", "请输入行情日期（YYYYMMDD格式）")
            return False
        
        # 验证日期格式
        if len(self.query_date.get()) != 8 or not self.query_date.get().isdigit():
            messagebox.showerror("错误", "行情日期格式错误，请输入YYYYMMDD格式")
            return False
        
        if not self.valuation_date:
            messagebox.showerror("错误", "请选择估值表时间")
            return False
        
        return True
    
    def _start_process(self):
        """开始处理"""
        if not self._validate_inputs():
            return
        
        self._clear_log()
        self._log("=" * 50)
        self._log("开始处理虚拟净值数据...", "info")
        self._log("=" * 50)
        
        # 在新线程中执行处理
        thread = threading.Thread(target=self._process_thread)
        thread.daemon = True
        thread.start()
    
    def _process_thread(self):
        """处理线程"""
        try:
            result = process_valuation_data(
                self.excel_folder.get(),
                self.query_date.get(),
                self.valuation_date,
                progress_callback=self._update_progress,
                log_callback=self._log
            )
            
            if result:
                messagebox.showinfo("成功", "数据处理完成！")
            else:
                messagebox.showerror("失败", "数据处理失败，请查看日志")
        
        except Exception as e:
            self._log(f"❌ 处理过程出错: {str(e)}", "error")
            messagebox.showerror("错误", f"处理过程出错: {str(e)}")


# ==================== 主程序入口 ====================
def main():
    root = tk.Tk()
    app = ValuationGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
